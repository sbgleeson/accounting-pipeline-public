from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from accounting_pipeline.config import (
    OUTPUT_COLUMNS,
    OUTPUT_XLSX_FILE,
    ProfileSettings,
    get_owner_buckets,
    load_accounts,
)
from accounting_pipeline.models import Account, StatementMetadata, Transaction, VenmoActivity
from accounting_pipeline.transforms.venmo_enrichment import deduplicate_activities
from accounting_pipeline.transforms.dedupe import get_sorted_rows
from accounting_pipeline.reports.income_transfer_check import (
    INCOME_ROUTING_HEADERS,
    INCOME_ROUTING_SUMMARY_HEADERS,
    build_income_routing_rows,
    build_income_routing_summary_rows,
)
from accounting_pipeline.reports.monthly_summary import (
    get_income_category_groups,
    get_reporting_months,
    get_summary_category_groups,
    populate_cash_flow_summary,
    populate_income_summary,
    populate_monthly_summary,
)
from accounting_pipeline.reports.reconciliation import (
    RECONCILIATION_HEADERS,
    build_reconciliation_rows,
)
from accounting_pipeline.reports.workbook_front_matter import (
    NEEDS_REVIEW_HEADERS,
    build_needs_review_rows,
    build_overview_metrics,
)
from accounting_pipeline.utils import decimal_to_number, parse_date


HIDDEN_TRANSACTION_COLUMNS = {
    "account_id",
    "account_type",
    "source_file",
    "raw_type",
    "details",
    "balance",
    "check_number",
    "transfer_group_id",
    "counterparty_account_id",
    "venmo_match_type",
    "venmo_id",
    "venmo_datetime",
    "venmo_from",
    "venmo_to",
    "venmo_source_file",
}

CATEGORY_COLORS = {
    "Housing": "EADCF8",
    "Food": "D9EAD3",
    "Auto + Transport": "D9EAF7",
    "Travel": "DDEBF7",
    "Health": "FCE4D6",
    "Shopping": "FFF2CC",
    "Entertainment": "E2F0D9",
    "Financial": "D9D2E9",
    "Transfers": "DAEEF3",
    "Savings": "C9DAF8",
    "Investing": "D0E0E3",
    "Giving": "FCE5CD",
    "Income": "C6E0B4",
    "Uncategorized": "F4CCCC",
}

TRANSACTION_COLUMN_WIDTHS = {
    "account_name": 25,
    "owner_bucket": 14,
    "transaction_date": 13,
    "post_date": 13,
    "description": 42,
    "canonical_merchant": 28,
    "amount": 12,
    "category": 34,
    "category_source": 24,
    "activity_type": 20,
    "memo": 32,
    "is_internal_transfer": 12,
    "venmo_match_status": 14,
    "venmo_note": 32,
}

VENMO_ACTIVITY_HEADERS = [
    "venmo_datetime",
    "venmo_date",
    "venmo_id",
    "activity_type",
    "status",
    "amount",
    "from_name",
    "to_name",
    "note",
    "funding_source",
    "destination",
    "source_file",
    "chase_link_status",
    "chase_match_type",
    "chase_post_date",
    "chase_amount",
    "chase_description",
    "chase_category",
    "chase_owner_bucket",
    "chase_source_file",
]

VENMO_ACTIVITY_COLUMN_WIDTHS = {
    "venmo_datetime": 20,
    "venmo_date": 13,
    "venmo_id": 22,
    "activity_type": 16,
    "status": 12,
    "amount": 12,
    "from_name": 20,
    "to_name": 20,
    "note": 40,
    "funding_source": 18,
    "destination": 24,
    "source_file": 26,
    "chase_link_status": 18,
    "chase_match_type": 16,
    "chase_post_date": 13,
    "chase_amount": 12,
    "chase_description": 40,
    "chase_category": 26,
    "chase_owner_bucket": 14,
    "chase_source_file": 26,
}

RECONCILIATION_COLUMN_WIDTHS = {
    "account_id": 10,
    "account_name": 18,
    "account_type": 13,
    "statement_start_date": 16,
    "statement_end_date": 16,
    "opening_balance_input": 18,
    "closing_balance_input": 18,
    "net_activity_in_period": 18,
    "expected_closing_balance": 20,
    "difference_to_statement": 20,
    "status": 10,
}

MONTHLY_SUMMARY_COLUMN_WIDTHS = {
    "category": 34,
    "total_spend": 14,
}

INCOME_ROUTING_COLUMN_WIDTHS = {
    "row_type": 24,
    "post_date": 11,
    "amount": 11,
    "routing_status": 28,
    "routing_note": 48,
    "category": 26,
    "account_name": 15,
    "owner_bucket": 11,
    "description": 28,
    "memo": 20,
    "canonical_merchant": 18,
    "venmo_from": 14,
    "venmo_to": 14,
    "venmo_note": 20,
}

INCOME_ROUTING_SUMMARY_COLUMN_WIDTHS = {
    "year": 10,
    "observed_income": 18,
    "income_in_destination": 20,
    "income_outside_destination": 22,
    "internal_transfers_into_destination": 24,
    "visibility_note": 52,
}

CATEGORIES_BUDGET_HEADERS = [
    "budget_label",
    "budget_level",
    "main_category",
    "sub_category",
    "combined_category",
    "reality_monthly",
    "monthly_target",
    "target_type",
    "owner_bucket",
    "review_status",
    "notes",
]

CATEGORIES_BUDGET_COLUMN_WIDTHS = {
    "budget_label": 34,
    "budget_level": 18,
    "main_category": 18,
    "sub_category": 28,
    "combined_category": 34,
    "reality_monthly": 15,
    "monthly_target": 15,
    "target_type": 12,
    "owner_bucket": 14,
    "review_status": 14,
    "notes": 52,
}

NEEDS_REVIEW_COLUMN_WIDTHS = {
    "review_reason": 28,
    "post_date": 13,
    "amount": 12,
    "account_name": 24,
    "description": 38,
    "category": 32,
    "category_source": 22,
    "owner_bucket": 14,
    "venmo_status": 14,
    "review_note": 42,
}

SUMMARY_TABLE_START_ROW = 7


def populate_overview_sheet(
    worksheet,
    metrics: list[tuple[str, object, str]],
    income_routing_enabled: bool,
) -> None:
    """Create a concise presentation and navigation surface."""
    worksheet.merge_cells("A1:I2")
    worksheet["A1"] = "Accounting Workbook Overview"
    worksheet.merge_cells("A3:I3")
    worksheet["A3"] = (
        "Start here for the loaded period, headline activity, review workload, and workbook navigation. "
        "Totals reflect only the accounts and files loaded into this profile."
    )

    card_columns = (("A", "C"), ("D", "F"), ("G", "I"))
    for index, (label, value, note) in enumerate(metrics):
        card_row = 5 + (index // 3) * 4
        start_column, end_column = card_columns[index % 3]
        worksheet.merge_cells(f"{start_column}{card_row}:{end_column}{card_row}")
        worksheet.merge_cells(f"{start_column}{card_row + 1}:{end_column}{card_row + 1}")
        worksheet.merge_cells(f"{start_column}{card_row + 2}:{end_column}{card_row + 2}")
        worksheet[f"{start_column}{card_row}"] = label
        worksheet[f"{start_column}{card_row + 1}"] = decimal_to_number(value) if isinstance(value, Decimal) else value
        worksheet[f"{start_column}{card_row + 2}"] = note
        if isinstance(value, Decimal):
            worksheet[f"{start_column}{card_row + 1}"].number_format = "$#,##0.00"

    guide_start_row = 18
    worksheet.merge_cells(start_row=guide_start_row, start_column=1, end_row=guide_start_row, end_column=9)
    worksheet.cell(guide_start_row, 1).value = "Workbook guide"
    worksheet.cell(guide_start_row + 1, 1).value = "Sheet"
    worksheet.cell(guide_start_row + 1, 3).value = "What it answers"
    guide_rows = [
        ("Needs Review", "What requires a person to classify, confirm, or investigate?"),
        ("Categories & Budget", "What categories and editable monthly targets drive the summaries?"),
        ("Spending Summary", "Where did money go after excluding transfers and card payments?"),
        ("Income Summary", "What income is visible in the loaded accounts, by source?"),
        ("Cash Flow Summary", "What moved into and out of cash accounts?"),
        ("transactions", "What normalized activity was loaded, and how was each row categorized?"),
        ("venmo_activity", "Which raw Venmo export rows linked to loaded Chase transactions?"),
        ("reconciliation", "Which statement periods reconcile to loaded transaction activity?"),
    ]
    if income_routing_enabled:
        guide_rows.append(
            (
                "Income Routing Review",
                "Where is income observed, and what internal transfers into the family destination are visible?",
            )
        )
    for row_number, (sheet_name, purpose) in enumerate(guide_rows, start=guide_start_row + 2):
        worksheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
        worksheet.merge_cells(start_row=row_number, start_column=3, end_row=row_number, end_column=9)
        worksheet.cell(row_number, 1).value = sheet_name
        worksheet.cell(row_number, 3).value = purpose


def populate_needs_review_sheet(worksheet, review_rows: list[list[object]]) -> None:
    """Create one compact exception list for human review."""
    worksheet.merge_cells("A1:J1")
    worksheet["A1"] = "Needs Review"
    worksheet.merge_cells("A2:J2")
    worksheet["A2"] = (
        "Items remain visible until their category, owner bucket, Venmo match, or statement coverage is resolved."
    )
    worksheet.append([])
    worksheet.append(NEEDS_REVIEW_HEADERS)
    for review_row in review_rows:
        worksheet.append(
            [decimal_to_number(value) if isinstance(value, Decimal) else value for value in review_row]
        )


def get_group_main_rows(
    category_groups: list[tuple[str, list[str]]],
    start_row: int = SUMMARY_TABLE_START_ROW,
) -> list[int]:
    """Return row numbers containing main summary categories."""
    rows = []
    row_number = start_row + 3
    for _main_category, category_names in category_groups:
        rows.append(row_number)
        row_number += len(category_names) + 1
    return rows


def populate_summary_front_matter(
    worksheet,
    title: str,
    subtitle: str,
    cards: list[tuple[str, str]],
) -> None:
    """Add a large title and four formula-backed summary cards."""
    title_end_column = get_column_letter(max(worksheet.max_column, 8))
    worksheet.merge_cells(f"A1:{title_end_column}1")
    worksheet["A1"] = title
    worksheet.merge_cells(f"A2:{title_end_column}2")
    worksheet["A2"] = subtitle
    card_ranges = (("A", "B"), ("C", "D"), ("E", "F"), ("G", "H"))
    for (label, formula), (start_column, end_column) in zip(cards, card_ranges):
        worksheet.merge_cells(f"{start_column}4:{end_column}4")
        worksheet.merge_cells(f"{start_column}5:{end_column}5")
        worksheet[f"{start_column}4"] = label
        worksheet[f"{start_column}5"] = formula


def get_main_category(category_name: str) -> str:
    """Return the main category from a combined category label."""
    return category_name.split(" – ", 1)[0]


def get_default_budget_target_type(main_category: str) -> str:
    """Return the default target behavior for a category budget row."""
    if main_category in {"Income", "Savings", "Investing"}:
        return "min"
    if main_category in {"Transfers", "Uncategorized"}:
        return "review"
    return "max"


def populate_categories_budget_sheet(
    worksheet,
    category_rows: list[dict[str, str]],
    budget_targets: list[dict[str, str]] | None = None,
) -> None:
    """Write the editable category-and-budget planning sheet."""
    targets_by_label = {
        row["budget_label"]: row
        for row in (budget_targets or [])
    }
    worksheet.append(CATEGORIES_BUDGET_HEADERS)
    human_headers = [
        "Category / budget label",
        "Budget level",
        "Main category",
        "Subcategory",
        "Combined category",
        "Reality",
        "Monthly target",
        "Target type",
        "Owner bucket",
        "Review status",
        "Notes",
    ]
    for column_number, header in enumerate(human_headers, start=1):
        worksheet.cell(1, column_number).value = header
    seen_main_categories: set[str] = set()
    written_labels: set[str] = set()

    for category_row in category_rows:
        main_category = category_row["main_category"]
        if main_category not in seen_main_categories:
            seen_main_categories.add(main_category)
            target = targets_by_label.get(main_category, {})
            worksheet.append(
                [
                    main_category,
                    "main_category",
                    main_category,
                    "",
                    "",
                    float(target["reality_monthly"]) if target.get("reality_monthly") else "",
                    float(target["monthly_target"]) if target.get("monthly_target") else "",
                    target.get("target_type") or get_default_budget_target_type(main_category),
                    target.get("owner_bucket", ""),
                    target.get("review_status", ""),
                    target.get("notes", ""),
                ]
            )
            written_labels.add(main_category)

        combined = category_row["combined"]
        target = targets_by_label.get(combined, {})
        worksheet.append(
            [
                combined,
                "combined_category",
                main_category,
                category_row["sub_category"],
                combined,
                float(target["reality_monthly"]) if target.get("reality_monthly") else "",
                float(target["monthly_target"]) if target.get("monthly_target") else "",
                target.get("target_type") or get_default_budget_target_type(main_category),
                target.get("owner_bucket", ""),
                target.get("review_status", ""),
                target.get("notes", ""),
            ]
        )
        written_labels.add(combined)

    for label, target in targets_by_label.items():
        if label in written_labels:
            continue
        worksheet.append(
            [
                label,
                "summary_target",
                label.split(" – ", 1)[0],
                "",
                "",
                float(target["reality_monthly"]) if target.get("reality_monthly") else "",
                float(target["monthly_target"]),
                target.get("target_type") or "review",
                target.get("owner_bucket", ""),
                target.get("review_status", ""),
                target.get("notes", ""),
            ]
        )


def build_venmo_activity_links(rows: list[Transaction]) -> dict[str, Transaction]:
    """Return the Chase row linked to each Venmo activity ID."""
    links: dict[str, Transaction] = {}
    for row in rows:
        if not row.venmo_id:
            continue
        for venmo_id in str(row.venmo_id).split(";"):
            normalized_id = venmo_id.strip()
            if normalized_id and normalized_id not in links:
                links[normalized_id] = row
    return links


def populate_venmo_activity_sheet(
    worksheet,
    activities: list[VenmoActivity],
    rows: list[Transaction],
) -> None:
    """Write raw Venmo export activity with any linked Chase transaction details."""
    worksheet.append(VENMO_ACTIVITY_HEADERS)
    links_by_venmo_id = build_venmo_activity_links(rows)

    for activity in sorted(
        deduplicate_activities(activities),
        key=lambda item: (item.datetime, item.venmo_id),
        reverse=True,
    ):
        linked_row = links_by_venmo_id.get(activity.venmo_id)
        venmo_datetime = datetime.strptime(activity.datetime, "%Y-%m-%dT%H:%M:%S")
        worksheet.append(
            [
                venmo_datetime,
                venmo_datetime.date(),
                activity.venmo_id,
                activity.activity_type,
                activity.status,
                decimal_to_number(activity.amount),
                activity.from_name,
                activity.to_name,
                activity.note,
                activity.funding_source,
                activity.destination,
                activity.source_file,
                linked_row.venmo_match_status if linked_row else "not_linked",
                linked_row.venmo_match_type if linked_row else "",
                parse_date(linked_row.post_date) if linked_row else "",
                decimal_to_number(linked_row.amount) if linked_row else "",
                linked_row.description if linked_row else "",
                linked_row.category if linked_row else "",
                linked_row.owner_bucket if linked_row else "",
                linked_row.source_file if linked_row else "",
            ]
        )


def write_excel_output(
    rows: list[Transaction],
    category_rows: list[dict[str, str]],
    statement_metadata: dict[str, list[StatementMetadata]],
    accounts: list[Account] | None = None,
    profile_settings: ProfileSettings = ProfileSettings(),
    budget_targets: list[dict[str, str]] | None = None,
    venmo_activities: list[VenmoActivity] | None = None,
    output_file: Path = OUTPUT_XLSX_FILE,
) -> None:
    """Write an Excel workbook with transactions, reconciliation, and summaries."""
    workbook = Workbook()
    sorted_rows = get_sorted_rows(rows)
    configured_accounts = accounts if accounts is not None else load_accounts()
    owner_buckets = get_owner_buckets(configured_accounts)
    review_rows = build_needs_review_rows(sorted_rows, configured_accounts, statement_metadata)

    overview_ws = workbook.active
    overview_ws.title = "Overview"
    populate_overview_sheet(
        overview_ws,
        build_overview_metrics(sorted_rows, configured_accounts, len(review_rows)),
        profile_settings.enable_income_routing_review,
    )

    needs_review_ws = workbook.create_sheet("Needs Review")
    populate_needs_review_sheet(needs_review_ws, review_rows)

    categories_ws = workbook.create_sheet("categories")
    categories_ws.append(["main_category", "sub_category", "combined"])
    for category_row in category_rows:
        categories_ws.append(
            [category_row["main_category"], category_row["sub_category"], category_row["combined"]]
        )

    last_category_row = len(category_rows) + 1
    workbook.defined_names.add(
        DefinedName("AllCategories", attr_text=f"'categories'!$C$2:$C${last_category_row}")
    )

    categories_budget_ws = workbook.create_sheet("Categories & Budget")
    populate_categories_budget_sheet(categories_budget_ws, category_rows, budget_targets)

    accounts_ws = workbook.create_sheet("accounts")
    accounts_ws.append(["account_id", "account_name", "account_type", "default_bucket"])
    for account in configured_accounts:
        accounts_ws.append(
            [account.account_id, account.account_name, account.account_type, account.default_bucket]
        )
    accounts_ws["F1"] = "owner_bucket_options"
    for row_number, bucket in enumerate(owner_buckets, start=2):
        accounts_ws[f"F{row_number}"] = bucket
    workbook.defined_names.add(
        DefinedName("AllOwnerBuckets", attr_text=f"'accounts'!$F$2:$F${len(owner_buckets) + 1}")
    )

    last_categories_budget_row = categories_budget_ws.max_row
    budget_level_validation = DataValidation(
        type="list",
        formula1='"main_category,combined_category,summary_target"',
        allow_blank=False,
    )
    categories_budget_ws.add_data_validation(budget_level_validation)
    budget_level_validation.add(f"B2:B{last_categories_budget_row}")

    budget_target_type_validation = DataValidation(
        type="list",
        formula1='"max,min,exact,review"',
        allow_blank=False,
    )
    categories_budget_ws.add_data_validation(budget_target_type_validation)
    budget_target_type_validation.add(f"H2:H{last_categories_budget_row}")

    budget_bucket_validation = DataValidation(type="list", formula1="=AllOwnerBuckets", allow_blank=True)
    categories_budget_ws.add_data_validation(budget_bucket_validation)
    budget_bucket_validation.add(f"I2:I{last_categories_budget_row}")

    budget_review_status_validation = DataValidation(
        type="list",
        formula1='"draft,agreed,needs_review,one_time_or_irregular"',
        allow_blank=True,
    )
    categories_budget_ws.add_data_validation(budget_review_status_validation)
    budget_review_status_validation.add(f"J2:J{last_categories_budget_row}")

    transactions_ws = workbook.create_sheet("transactions")
    transactions_ws.append(OUTPUT_COLUMNS)
    for row in sorted_rows:
        transactions_ws.append(
            [
                row.account_id,
                row.account_name,
                row.account_type,
                row.owner_bucket,
                row.source_file,
                parse_date(row.transaction_date),
                parse_date(row.post_date),
                row.description,
                row.canonical_merchant,
                decimal_to_number(row.amount),
                row.raw_type,
                row.details,
                decimal_to_number(row.balance),
                row.category,
                row.category_source,
                row.activity_type,
                row.memo,
                row.check_number,
                "true" if row.is_internal_transfer else "false",
                row.transfer_group_id,
                row.counterparty_account_id,
                row.venmo_match_status,
                row.venmo_match_type,
                row.venmo_id,
                row.venmo_datetime,
                row.venmo_from,
                row.venmo_to,
                row.venmo_note,
                row.venmo_source_file,
            ]
        )

    venmo_activity_ws = workbook.create_sheet("venmo_activity")
    populate_venmo_activity_sheet(venmo_activity_ws, venmo_activities or [], sorted_rows)

    category_column_index = OUTPUT_COLUMNS.index("category") + 1
    category_column_letter = get_column_letter(category_column_index)
    owner_bucket_column_index = OUTPUT_COLUMNS.index("owner_bucket") + 1
    owner_bucket_column_letter = get_column_letter(owner_bucket_column_index)
    last_transaction_row = len(sorted_rows) + 1
    amount_column_letter = get_column_letter(OUTPUT_COLUMNS.index("amount") + 1)
    account_column_letter = get_column_letter(OUTPUT_COLUMNS.index("account_id") + 1)
    account_type_column_letter = get_column_letter(OUTPUT_COLUMNS.index("account_type") + 1)
    post_date_column_letter = get_column_letter(OUTPUT_COLUMNS.index("post_date") + 1)
    bucket_column_letter = get_column_letter(OUTPUT_COLUMNS.index("owner_bucket") + 1)

    category_validation = DataValidation(type="list", formula1="=AllCategories", allow_blank=True)
    transactions_ws.add_data_validation(category_validation)
    if last_transaction_row >= 2:
        category_validation.add(f"{category_column_letter}2:{category_column_letter}{last_transaction_row}")

    bucket_validation = DataValidation(type="list", formula1="=AllOwnerBuckets", allow_blank=False)
    transactions_ws.add_data_validation(bucket_validation)
    if last_transaction_row >= 2:
        bucket_validation.add(f"{owner_bucket_column_letter}2:{owner_bucket_column_letter}{last_transaction_row}")

    reconciliation_ws = workbook.create_sheet("reconciliation")
    reconciliation_ws.append(RECONCILIATION_HEADERS)

    amount_range = f"transactions!${amount_column_letter}$2:${amount_column_letter}${last_transaction_row}"
    account_range = f"transactions!${account_column_letter}$2:${account_column_letter}${last_transaction_row}"
    account_type_range = f"transactions!${account_type_column_letter}$2:${account_type_column_letter}${last_transaction_row}"
    post_date_range = f"transactions!${post_date_column_letter}$2:${post_date_column_letter}${last_transaction_row}"
    bucket_range = f"transactions!${bucket_column_letter}$2:${bucket_column_letter}${last_transaction_row}"
    category_range = f"transactions!${category_column_letter}$2:${category_column_letter}${last_transaction_row}"

    for row in build_reconciliation_rows(
        statement_metadata,
        amount_range,
        account_range,
        post_date_range,
        configured_accounts,
    ):
        reconciliation_ws.append(row)

    category_groups = get_summary_category_groups(category_rows)
    income_category_groups = get_income_category_groups(category_rows)
    reporting_months = get_reporting_months([parse_date(row.post_date) for row in sorted_rows])
    cash_flow_summary_ws = workbook.create_sheet("Cash Flow Summary")
    populate_cash_flow_summary(
        cash_flow_summary_ws,
        reporting_months,
        amount_range,
        category_range,
        account_type_range,
        post_date_range,
        start_row=SUMMARY_TABLE_START_ROW,
    )

    income_summary_ws = workbook.create_sheet("Income Summary")
    populate_income_summary(
        income_summary_ws,
        income_category_groups,
        reporting_months,
        amount_range,
        category_range,
        account_type_range,
        post_date_range,
        start_row=SUMMARY_TABLE_START_ROW,
    )

    income_routing_ws = None
    income_routing_summary_start_row = None
    income_routing_summary_header_row = None
    if profile_settings.enable_income_routing_review:
        income_routing_ws = workbook.create_sheet("Income Routing Review")
        income_routing_ws.merge_cells("A1:N1")
        income_routing_ws["A1"] = "Income Routing Review"
        income_routing_ws.merge_cells("A2:N2")
        income_routing_ws["A2"] = (
            "Family-specific diagnostic: observed income and internal transfers are shown separately. "
            "The workbook does not claim that a transfer belongs to a specific income deposit."
        )
        income_routing_ws.append([])
        income_routing_ws.append(INCOME_ROUTING_HEADERS)
        for report_row in build_income_routing_rows(sorted_rows, profile_settings):
            income_routing_ws.append(
                [
                    decimal_to_number(value) if isinstance(value, Decimal) else value
                    for value in report_row
                ]
            )
        income_routing_summary_start_row = income_routing_ws.max_row + 3
        income_routing_ws.cell(row=income_routing_summary_start_row, column=1).value = (
            "Yearly visibility summary"
        )
        income_routing_ws.cell(row=income_routing_summary_start_row + 1, column=1).value = (
            INCOME_ROUTING_SUMMARY_HEADERS[0]
        )
        for column_number, header in enumerate(INCOME_ROUTING_SUMMARY_HEADERS[1:], start=2):
            income_routing_ws.cell(
                row=income_routing_summary_start_row + 1,
                column=column_number,
            ).value = header
        income_routing_summary_header_row = income_routing_summary_start_row + 1
        for report_row in build_income_routing_summary_rows(sorted_rows, profile_settings):
            income_routing_ws.append(
                [
                    decimal_to_number(value) if isinstance(value, Decimal) else value
                    for value in report_row
                ]
            )

    monthly_summary_ws = workbook.create_sheet("Spending Summary")
    populate_monthly_summary(
        monthly_summary_ws,
        category_groups,
        reporting_months,
        amount_range,
        category_range,
        bucket_range,
        post_date_range,
        owner_buckets,
        profile_settings.credit_bucket,
        start_row=SUMMARY_TABLE_START_ROW,
    )

    latest_month_column = len(reporting_months) + 1
    total_column = len(reporting_months) + 2
    target_column = total_column + 1
    target_type_column = total_column + 2
    monthly_variance_column = total_column + 3
    ytd_actual_column = total_column + 4
    ytd_target_column = total_column + 5
    ytd_variance_column = total_column + 6
    average_column = total_column + 7
    average_variance_column = total_column + 8
    latest_month_letter = get_column_letter(latest_month_column)
    target_letter = get_column_letter(target_column)
    average_letter = get_column_letter(average_column)
    ytd_variance_letter = get_column_letter(ytd_variance_column)

    spending_main_rows = get_group_main_rows(category_groups)
    income_main_rows = get_group_main_rows(income_category_groups)
    spending_latest_cells = ",".join(f"{latest_month_letter}{row}" for row in spending_main_rows)
    spending_average_cells = ",".join(f"{average_letter}{row}" for row in spending_main_rows)
    spending_target_cells = ",".join(f"{target_letter}{row}" for row in spending_main_rows)
    spending_review_count = "+".join(
        f'COUNTIF({ytd_variance_letter}{row},"<0")' for row in spending_main_rows
    )
    populate_summary_front_matter(
        monthly_summary_ws,
        "Spending Summary",
        "Latest-month results and budget signals first; expand grouped rows for subcategory and historical detail.",
        [
            ("Latest month spending", f"=SUM({spending_latest_cells})"),
            ("Monthly average", f"=SUM({spending_average_cells})"),
            ("Configured targets", f"=SUM({spending_target_cells})"),
            ("Categories to review", f"={spending_review_count or '0'}"),
        ],
    )

    income_latest_cells = ",".join(f"{latest_month_letter}{row}" for row in income_main_rows)
    income_average_cells = ",".join(f"{average_letter}{row}" for row in income_main_rows)
    income_target_cells = ",".join(f"{target_letter}{row}" for row in income_main_rows)
    income_review_count = "+".join(
        f'COUNTIF({ytd_variance_letter}{row},"<0")' for row in income_main_rows
    )
    populate_summary_front_matter(
        income_summary_ws,
        "Income Summary",
        "Observed income in loaded accounts, with the latest month and configured minimum targets emphasized.",
        [
            ("Latest month income", f"=SUM({income_latest_cells})"),
            ("Monthly average", f"=SUM({income_average_cells})"),
            ("Expected monthly income", f"=SUM({income_target_cells or '0'})"),
            ("Income targets to review", f"={income_review_count or '0'}"),
        ],
    )

    cash_first_data_row = SUMMARY_TABLE_START_ROW + 3
    cash_in_row = cash_first_data_row
    cash_out_row = cash_first_data_row + 1
    cash_net_row = cash_first_data_row + 2
    cash_ytd_column = total_column + 1
    cash_ytd_letter = get_column_letter(cash_ytd_column)
    populate_summary_front_matter(
        cash_flow_summary_ws,
        "Cash Flow Summary",
        "Cash-account movement for the latest loaded month; internal transfers and card payments remain separate below.",
        [
            ("Latest cash in", f"={latest_month_letter}{cash_in_row}"),
            ("Latest cash out", f"={latest_month_letter}{cash_out_row}"),
            ("Latest net cash flow", f"={latest_month_letter}{cash_net_row}"),
            ("YTD net cash flow", f"={cash_ytd_letter}{cash_net_row}"),
        ],
    )

    for row in transactions_ws.iter_rows(min_row=2, max_row=last_transaction_row, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "mm/dd/yyyy"
    amount_column_index = OUTPUT_COLUMNS.index("amount") + 1
    for row in transactions_ws.iter_rows(
        min_row=2,
        max_row=last_transaction_row,
        min_col=amount_column_index,
        max_col=amount_column_index,
    ):
        for cell in row:
            cell.number_format = "$#,##0.00"
    balance_column_index = OUTPUT_COLUMNS.index("balance") + 1
    for row in transactions_ws.iter_rows(
        min_row=2,
        max_row=last_transaction_row,
        min_col=balance_column_index,
        max_col=balance_column_index,
    ):
        for cell in row:
            cell.number_format = "$#,##0.00"

    for row_number in range(2, reconciliation_ws.max_row + 1):
        reconciliation_ws[f"D{row_number}"].number_format = "mm/dd/yyyy"
        reconciliation_ws[f"E{row_number}"].number_format = "mm/dd/yyyy"
        for column_letter in ("F", "G", "H", "I", "J"):
            reconciliation_ws[f"{column_letter}{row_number}"].number_format = "$#,##0.00"

    for row in monthly_summary_ws.iter_rows(
        min_row=SUMMARY_TABLE_START_ROW + 1,
        max_row=SUMMARY_TABLE_START_ROW + 2,
        min_col=2,
        max_col=len(reporting_months) + 1,
    ):
        for cell in row:
            cell.number_format = "mm/dd/yyyy"
    for row in monthly_summary_ws.iter_rows(
        min_row=SUMMARY_TABLE_START_ROW + 3,
        max_row=monthly_summary_ws.max_row,
        min_col=2,
        max_col=len(reporting_months) + 2,
    ):
        for cell in row:
            cell.number_format = "$#,##0.00"
    for row in income_summary_ws.iter_rows(
        min_row=SUMMARY_TABLE_START_ROW + 1,
        max_row=SUMMARY_TABLE_START_ROW + 2,
        min_col=2,
        max_col=len(reporting_months) + 1,
    ):
        for cell in row:
            cell.number_format = "mm/dd/yyyy"
    for row in income_summary_ws.iter_rows(
        min_row=SUMMARY_TABLE_START_ROW + 3,
        max_row=income_summary_ws.max_row,
        min_col=2,
        max_col=len(reporting_months) + 2,
    ):
        for cell in row:
            cell.number_format = "$#,##0.00"
    for row_number in range(5, needs_review_ws.max_row + 1):
        needs_review_ws[f"B{row_number}"].number_format = "mm/dd/yyyy"
        needs_review_ws[f"C{row_number}"].number_format = "$#,##0.00"
    if income_routing_ws is not None:
        for row_number in range(5, income_routing_summary_start_row):
            income_routing_ws[f"B{row_number}"].number_format = "mm/dd/yyyy"
            income_routing_ws[f"C{row_number}"].number_format = "$#,##0.00"
        for row_number in range(
            income_routing_summary_header_row + 1,
            income_routing_ws.max_row + 1,
        ):
            for column_letter in ("B", "C", "D", "E"):
                income_routing_ws[f"{column_letter}{row_number}"].number_format = "$#,##0.00"
    for row in cash_flow_summary_ws.iter_rows(
        min_row=SUMMARY_TABLE_START_ROW + 1,
        max_row=SUMMARY_TABLE_START_ROW + 2,
        min_col=2,
        max_col=len(reporting_months) + 1,
    ):
        for cell in row:
            cell.number_format = "mm/dd/yyyy"
    for row in cash_flow_summary_ws.iter_rows(
        min_row=SUMMARY_TABLE_START_ROW + 3,
        max_row=cash_flow_summary_ws.max_row,
        min_col=2,
        max_col=cash_flow_summary_ws.max_column,
    ):
        for cell in row:
            cell.number_format = "$#,##0.00"
    for worksheet in (monthly_summary_ws, income_summary_ws, cash_flow_summary_ws):
        for column_number in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=SUMMARY_TABLE_START_ROW, column=column_number).value
            if header in {
                "Monthly target",
                "Monthly variance",
                "YTD actual",
                "YTD target",
                "YTD variance",
                "Monthly average",
                "Average variance",
                "Variance",
            }:
                for row in worksheet.iter_rows(
                    min_row=SUMMARY_TABLE_START_ROW + 3,
                    max_row=worksheet.max_row,
                    min_col=column_number,
                    max_col=column_number,
                ):
                    for cell in row:
                        cell.number_format = "$#,##0.00"
    for column_name in ("reality_monthly", "monthly_target"):
        for row in categories_budget_ws.iter_rows(
            min_row=2,
            max_row=categories_budget_ws.max_row,
            min_col=CATEGORIES_BUDGET_HEADERS.index(column_name) + 1,
            max_col=CATEGORIES_BUDGET_HEADERS.index(column_name) + 1,
        ):
            for cell in row:
                cell.number_format = "$#,##0.00"

    sheet_header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    primary_budget_fill = PatternFill(fill_type="solid", fgColor="F4B183")
    reference_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    warning_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    neutral_fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
    summary_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    attention_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    bold_font = Font(bold=True, color="FFFFFF")
    dark_bold_font = Font(bold=True)
    title_font = Font(bold=True, color="FFFFFF", size=24)
    subtitle_font = Font(color="44546A", italic=True, size=13)
    metric_value_font = Font(bold=True, color="1F1F1F", size=18)
    thin_gray_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    cash_flow_net_font = Font(color="000000")
    cash_flow_net_negative_font = Font(color="C00000")
    budget_variance_met_font = Font(color="000000")
    budget_variance_missed_font = Font(color="C00000")
    category_fills = {
        category_name: PatternFill(fill_type="solid", fgColor=color)
        for category_name, color in CATEGORY_COLORS.items()
    }

    for worksheet in (
        categories_ws,
        categories_budget_ws,
        accounts_ws,
        transactions_ws,
        venmo_activity_ws,
        reconciliation_ws,
    ):
        for cell in worksheet[1]:
            cell.fill = sheet_header_fill
            cell.font = bold_font

    overview_ws["A1"].fill = sheet_header_fill
    overview_ws["A1"].font = title_font
    overview_ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    overview_ws["A3"].font = subtitle_font
    overview_ws["A3"].alignment = Alignment(wrap_text=True)
    for index in range(9):
        card_row = 5 + (index // 3) * 4
        start_column = ("A", "D", "G")[index % 3]
        end_column = ("C", "F", "I")[index % 3]
        for row_number in range(card_row, card_row + 3):
            for row in overview_ws[f"{start_column}{row_number}:{end_column}{row_number}"]:
                for cell in row:
                    cell.fill = summary_fill
                    cell.border = thin_gray_border
        overview_ws[f"{start_column}{card_row}"].font = Font(bold=True, size=12)
        overview_ws[f"{start_column}{card_row + 1}"].font = metric_value_font
        overview_ws[f"{start_column}{card_row + 2}"].font = Font(
            color="666666",
            italic=True,
            size=11,
        )
        overview_ws[f"{start_column}{card_row + 2}"].alignment = Alignment(wrap_text=True)
    overview_ws["A18"].fill = sheet_header_fill
    overview_ws["A18"].font = bold_font
    for cell in (overview_ws["A19"], overview_ws["C19"]):
        cell.fill = summary_fill
        cell.font = dark_bold_font
    for row_number in range(20, overview_ws.max_row + 1):
        overview_ws[f"A{row_number}"].font = Font(bold=True, size=11)
        overview_ws[f"C{row_number}"].font = Font(size=11)
        overview_ws[f"C{row_number}"].alignment = Alignment(wrap_text=True)
        overview_ws.row_dimensions[row_number].height = 21

    for summary_ws in (monthly_summary_ws, income_summary_ws, cash_flow_summary_ws):
        summary_ws.sheet_view.showGridLines = False
        summary_ws.sheet_view.zoomScale = 110
        summary_ws["A1"].fill = sheet_header_fill
        summary_ws["A1"].font = title_font
        summary_ws["A2"].font = subtitle_font
        summary_ws["A2"].alignment = Alignment(wrap_text=True)
        summary_ws.row_dimensions[1].height = 28
        summary_ws.row_dimensions[2].height = 30
        for start_column, end_column in (("A", "B"), ("C", "D"), ("E", "F"), ("G", "H")):
            for row_number in (4, 5):
                for row in summary_ws[f"{start_column}{row_number}:{end_column}{row_number}"]:
                    for cell in row:
                        cell.fill = summary_fill
                        cell.border = thin_gray_border
            summary_ws[f"{start_column}4"].font = dark_bold_font
            summary_ws[f"{start_column}5"].font = metric_value_font
            summary_ws[f"{start_column}5"].number_format = "$#,##0.00"
        if summary_ws in (monthly_summary_ws, income_summary_ws):
            summary_ws["G5"].number_format = "0"
        for cell in summary_ws[SUMMARY_TABLE_START_ROW]:
            if cell.value is not None:
                cell.fill = sheet_header_fill
                cell.font = bold_font

    needs_review_ws["A1"].fill = sheet_header_fill
    needs_review_ws["A1"].font = title_font
    needs_review_ws["A2"].font = subtitle_font
    needs_review_ws["A2"].alignment = Alignment(wrap_text=True)
    for cell in needs_review_ws[4]:
        cell.fill = sheet_header_fill
        cell.font = bold_font
    for row_number in range(5, needs_review_ws.max_row + 1):
        reason = str(needs_review_ws[f"A{row_number}"].value or "")
        if "Uncategorized" in reason or "Owner bucket" in reason:
            fill = warning_fill
        elif "Statement metadata" in reason:
            fill = neutral_fill
        else:
            fill = attention_fill
        for cell in needs_review_ws[row_number]:
            cell.fill = fill
        needs_review_ws[f"J{row_number}"].alignment = Alignment(wrap_text=True)

    if income_routing_ws is not None:
        income_routing_ws["A1"].fill = sheet_header_fill
        income_routing_ws["A1"].font = title_font
        income_routing_ws["A2"].font = subtitle_font
        income_routing_ws["A2"].alignment = Alignment(wrap_text=True)
        for cell in income_routing_ws[4]:
            cell.fill = sheet_header_fill
            cell.font = bold_font
        for row_number in range(5, income_routing_summary_start_row):
            row_type = income_routing_ws[f"A{row_number}"].value
            fill = reference_fill if row_type == "Income observed" else summary_fill
            for cell in income_routing_ws[row_number]:
                cell.fill = fill
            income_routing_ws[f"E{row_number}"].alignment = Alignment(wrap_text=True)
        income_routing_ws[f"A{income_routing_summary_start_row}"].fill = summary_fill
        income_routing_ws[f"A{income_routing_summary_start_row}"].font = dark_bold_font
        for cell in income_routing_ws[income_routing_summary_header_row]:
            if cell.value is not None:
                cell.fill = sheet_header_fill
                cell.font = bold_font
        for row_number in range(
            income_routing_summary_header_row + 1,
            income_routing_ws.max_row + 1,
        ):
            income_routing_ws[f"F{row_number}"].alignment = Alignment(wrap_text=True)

    for row_number in range(2, len(configured_accounts) + 2):
        accounts_ws[f"D{row_number}"].fill = editable_fill
        reconciliation_ws[f"D{row_number}"].fill = editable_fill
        reconciliation_ws[f"E{row_number}"].fill = editable_fill
        reconciliation_ws[f"F{row_number}"].fill = editable_fill
        reconciliation_ws[f"G{row_number}"].fill = editable_fill

    for row_number in range(2, categories_budget_ws.max_row + 1):
        main_category = categories_budget_ws[f"C{row_number}"].value
        if main_category in category_fills:
            categories_budget_ws[f"A{row_number}"].fill = category_fills[str(main_category)]
        for column_letter in ("F", "G", "H", "I", "J", "K"):
            categories_budget_ws[f"{column_letter}{row_number}"].fill = editable_fill
        if categories_budget_ws[f"B{row_number}"].value == "main_category":
            categories_budget_ws[f"G{row_number}"].fill = primary_budget_fill
            categories_budget_ws[f"H{row_number}"].fill = primary_budget_fill

    if last_transaction_row >= 2:
        for row_number in range(2, last_transaction_row + 1):
            transactions_ws[f"{owner_bucket_column_letter}{row_number}"].fill = editable_fill
            transactions_ws[f"{category_column_letter}{row_number}"].fill = editable_fill
            internal_transfer_column_letter = get_column_letter(OUTPUT_COLUMNS.index("is_internal_transfer") + 1)
            if transactions_ws[f"{internal_transfer_column_letter}{row_number}"].value == "true":
                for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
                    transactions_ws[f"{get_column_letter(column_index)}{row_number}"].fill = reference_fill
            category_value = transactions_ws[f"{category_column_letter}{row_number}"].value
            if category_value:
                main_category = get_main_category(str(category_value))
                if main_category in category_fills:
                    transactions_ws[f"{category_column_letter}{row_number}"].fill = category_fills[main_category]
            if transactions_ws[f"{category_column_letter}{row_number}"].value == "Uncategorized – Needs Review":
                transactions_ws[f"{category_column_letter}{row_number}"].fill = warning_fill
            if transactions_ws[f"{owner_bucket_column_letter}{row_number}"].value == "Needs Review":
                transactions_ws[f"{owner_bucket_column_letter}{row_number}"].fill = warning_fill

    monthly_summary_main_rows = []
    row_number = SUMMARY_TABLE_START_ROW + 3
    for main_category, category_names in category_groups:
        monthly_summary_main_rows.append((row_number, main_category))
        row_number += len(category_names) + 1
    bucket_header_row = row_number + 1
    for row_number, main_category in monthly_summary_main_rows:
        for cell in monthly_summary_ws[row_number]:
            cell.fill = category_fills.get(main_category, summary_fill)
            cell.font = dark_bold_font
    row_number = SUMMARY_TABLE_START_ROW + 3
    for _main_category, category_names in category_groups:
        monthly_summary_ws.row_dimensions[row_number].collapsed = True
        for child_row in range(row_number + 1, row_number + len(category_names) + 1):
            monthly_summary_ws.row_dimensions[child_row].hidden = True
            monthly_summary_ws.row_dimensions[child_row].outlineLevel = 1
        row_number += len(category_names) + 1
    for cell in monthly_summary_ws[bucket_header_row]:
        cell.fill = summary_fill
        cell.font = dark_bold_font
    income_summary_main_rows = []
    row_number = SUMMARY_TABLE_START_ROW + 3
    for main_category, category_names in income_category_groups:
        income_summary_main_rows.append((row_number, main_category))
        row_number += len(category_names) + 1
    for row_number, main_category in income_summary_main_rows:
        for cell in income_summary_ws[row_number]:
            cell.fill = category_fills.get(main_category, summary_fill)
            cell.font = dark_bold_font
    row_number = SUMMARY_TABLE_START_ROW + 3
    for _main_category, category_names in income_category_groups:
        income_summary_ws.row_dimensions[row_number].collapsed = True
        for child_row in range(row_number + 1, row_number + len(category_names) + 1):
            income_summary_ws.row_dimensions[child_row].hidden = True
            income_summary_ws.row_dimensions[child_row].outlineLevel = 1
        row_number += len(category_names) + 1
    for row_number in (
        SUMMARY_TABLE_START_ROW + 7,
        SUMMARY_TABLE_START_ROW + 12,
    ):
        for cell in cash_flow_summary_ws[row_number]:
            cell.fill = summary_fill
            cell.font = dark_bold_font
    cash_flow_net_row = SUMMARY_TABLE_START_ROW + 5
    cash_flow_value_columns = (
        f"B{cash_flow_net_row}:{get_column_letter(cash_flow_summary_ws.max_column)}{cash_flow_net_row}"
    )
    cash_flow_summary_ws.conditional_formatting.add(
        cash_flow_value_columns,
        CellIsRule(operator="lessThan", formula=["0"], font=cash_flow_net_negative_font),
    )
    cash_flow_summary_ws.conditional_formatting.add(
        cash_flow_value_columns,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=cash_flow_net_font),
    )
    for row_number in range(SUMMARY_TABLE_START_ROW + 7, SUMMARY_TABLE_START_ROW + 12):
        cash_flow_summary_ws.row_dimensions[row_number].hidden = True
        cash_flow_summary_ws.row_dimensions[row_number].outlineLevel = 1
    cash_flow_summary_ws.row_dimensions[SUMMARY_TABLE_START_ROW + 7].collapsed = True

    variance_column_letters = (
        get_column_letter(monthly_variance_column),
        get_column_letter(ytd_variance_column),
        get_column_letter(average_variance_column),
    )
    for summary_ws in (monthly_summary_ws, income_summary_ws):
        for variance_column_letter in variance_column_letters:
            variance_range = (
                f"{variance_column_letter}{SUMMARY_TABLE_START_ROW + 3}:"
                f"{variance_column_letter}{summary_ws.max_row}"
            )
            summary_ws.conditional_formatting.add(
                variance_range,
                CellIsRule(operator="lessThan", formula=["0"], font=budget_variance_missed_font),
            )
            summary_ws.conditional_formatting.add(
                variance_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=budget_variance_met_font),
            )

    transactions_ws.freeze_panes = "A2"
    needs_review_ws.freeze_panes = "A5"
    categories_budget_ws.freeze_panes = "A2"
    categories_budget_ws.sheet_view.zoomScale = 100
    categories_budget_ws.sheet_view.showGridLines = False
    venmo_activity_ws.freeze_panes = "A2"
    reconciliation_ws.freeze_panes = "A2"
    cash_flow_summary_ws.freeze_panes = f"B{SUMMARY_TABLE_START_ROW + 3}"
    income_summary_ws.freeze_panes = f"B{SUMMARY_TABLE_START_ROW + 3}"
    if income_routing_ws is not None:
        income_routing_ws.freeze_panes = "F5"
    monthly_summary_ws.freeze_panes = f"B{SUMMARY_TABLE_START_ROW + 3}"
    for summary_ws in (cash_flow_summary_ws, income_summary_ws, monthly_summary_ws):
        summary_ws.row_dimensions[SUMMARY_TABLE_START_ROW + 1].hidden = True
        summary_ws.row_dimensions[SUMMARY_TABLE_START_ROW + 2].hidden = True
    categories_budget_ws.auto_filter.ref = (
        f"A1:{get_column_letter(categories_budget_ws.max_column)}{categories_budget_ws.max_row}"
    )
    needs_review_ws.auto_filter.ref = f"A4:J{needs_review_ws.max_row}"
    if income_routing_ws is not None and income_routing_summary_start_row > 5:
        income_routing_ws.auto_filter.ref = (
            f"A4:N{income_routing_summary_start_row - 1}"
        )
    if last_transaction_row >= 1:
        transactions_ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{last_transaction_row}"

    for worksheet in (
        categories_ws,
        categories_budget_ws,
        accounts_ws,
        transactions_ws,
        venmo_activity_ws,
        reconciliation_ws,
    ):
        for column_cells in worksheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = max((len(value) for value in values), default=10)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 40)
    for column_letter in "ABCDEFGHI":
        overview_ws.column_dimensions[column_letter].width = 11
    overview_ws.row_dimensions[1].height = 32
    overview_ws.row_dimensions[2].height = 32
    overview_ws.row_dimensions[3].height = 38
    overview_ws.sheet_view.showGridLines = False
    overview_ws.sheet_view.zoomScale = 100
    overview_ws.sheet_properties.pageSetUpPr.fitToPage = True
    overview_ws.page_setup.orientation = "landscape"
    overview_ws.page_setup.fitToWidth = 1
    overview_ws.page_setup.fitToHeight = 1
    overview_ws.print_area = f"A1:I{overview_ws.max_row}"
    needs_review_ws.sheet_view.showGridLines = False
    needs_review_ws.row_dimensions[1].height = 28
    needs_review_ws.row_dimensions[2].height = 32
    for column_name, width in NEEDS_REVIEW_COLUMN_WIDTHS.items():
        needs_review_ws.column_dimensions[
            get_column_letter(NEEDS_REVIEW_HEADERS.index(column_name) + 1)
        ].width = width

    for column_name, width in TRANSACTION_COLUMN_WIDTHS.items():
        transactions_ws.column_dimensions[get_column_letter(OUTPUT_COLUMNS.index(column_name) + 1)].width = width

    for column_name, width in VENMO_ACTIVITY_COLUMN_WIDTHS.items():
        venmo_activity_ws.column_dimensions[
            get_column_letter(VENMO_ACTIVITY_HEADERS.index(column_name) + 1)
        ].width = width

    for column_name, width in RECONCILIATION_COLUMN_WIDTHS.items():
        reconciliation_ws.column_dimensions[
            get_column_letter(RECONCILIATION_HEADERS.index(column_name) + 1)
        ].width = width

    monthly_summary_ws.column_dimensions["A"].width = MONTHLY_SUMMARY_COLUMN_WIDTHS["category"]
    for column_number in range(2, monthly_summary_ws.max_column + 1):
        monthly_summary_ws.column_dimensions[get_column_letter(column_number)].width = 14
    income_summary_ws.column_dimensions["A"].width = MONTHLY_SUMMARY_COLUMN_WIDTHS["category"]
    for column_number in range(2, income_summary_ws.max_column + 1):
        income_summary_ws.column_dimensions[get_column_letter(column_number)].width = 14
    cash_flow_summary_ws.column_dimensions["A"].width = 32
    for column_number in range(2, cash_flow_summary_ws.max_column + 1):
        cash_flow_summary_ws.column_dimensions[get_column_letter(column_number)].width = 14
    for summary_ws in (monthly_summary_ws, income_summary_ws, cash_flow_summary_ws):
        for column_number in range(2, latest_month_column):
            column_letter = get_column_letter(column_number)
            summary_ws.column_dimensions[column_letter].hidden = True
            summary_ws.column_dimensions[column_letter].outlineLevel = 1
        if latest_month_column > 2:
            summary_ws.column_dimensions[get_column_letter(latest_month_column)].collapsed = True
        summary_ws.column_dimensions[get_column_letter(total_column)].hidden = True
        summary_ws.column_dimensions[get_column_letter(latest_month_column)].width = 16
    for summary_ws in (monthly_summary_ws, income_summary_ws):
        summary_ws.column_dimensions[get_column_letter(target_type_column)].hidden = True
        summary_ws.column_dimensions[get_column_letter(target_column)].width = 16
        summary_ws.column_dimensions[get_column_letter(average_column)].width = 16
        summary_ws.column_dimensions[get_column_letter(average_variance_column)].width = 16
    for header, width in {
        "Monthly variance": 16,
        "YTD actual": 14,
        "YTD target": 14,
        "YTD variance": 14,
        "Monthly average": 16,
        "Average variance": 16,
    }.items():
        for summary_ws in (monthly_summary_ws, income_summary_ws):
            for column_number in range(1, summary_ws.max_column + 1):
                if summary_ws.cell(row=SUMMARY_TABLE_START_ROW, column=column_number).value == header:
                    summary_ws.column_dimensions[get_column_letter(column_number)].width = width
                    break
    for column_name, width in CATEGORIES_BUDGET_COLUMN_WIDTHS.items():
        categories_budget_ws.column_dimensions[
            get_column_letter(CATEGORIES_BUDGET_HEADERS.index(column_name) + 1)
        ].width = width
    for column_letter in ("B", "C", "D", "E"):
        categories_budget_ws.column_dimensions[column_letter].hidden = True

    if income_routing_ws is not None:
        income_routing_ws.sheet_view.showGridLines = False
        income_routing_ws.row_dimensions[1].height = 28
        income_routing_ws.row_dimensions[2].height = 36
        for column_name, width in INCOME_ROUTING_COLUMN_WIDTHS.items():
            income_routing_ws.column_dimensions[
                get_column_letter(INCOME_ROUTING_HEADERS.index(column_name) + 1)
            ].width = width
        for column_name, width in INCOME_ROUTING_SUMMARY_COLUMN_WIDTHS.items():
            column_letter = get_column_letter(
                INCOME_ROUTING_SUMMARY_HEADERS.index(column_name) + 1
            )
            current_width = income_routing_ws.column_dimensions[column_letter].width or 10
            income_routing_ws.column_dimensions[column_letter].width = max(current_width, width)

    for row in venmo_activity_ws.iter_rows(min_row=2, max_row=venmo_activity_ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "mm/dd/yyyy"
    for column_name in ("amount", "chase_amount"):
        column_number = VENMO_ACTIVITY_HEADERS.index(column_name) + 1
        for row in venmo_activity_ws.iter_rows(
            min_row=2,
            max_row=venmo_activity_ws.max_row,
            min_col=column_number,
            max_col=column_number,
        ):
            for cell in row:
                cell.number_format = "$#,##0.00"
    chase_post_date_column = VENMO_ACTIVITY_HEADERS.index("chase_post_date") + 1
    for row in venmo_activity_ws.iter_rows(
        min_row=2,
        max_row=venmo_activity_ws.max_row,
        min_col=chase_post_date_column,
        max_col=chase_post_date_column,
    ):
        for cell in row:
            cell.number_format = "mm/dd/yyyy"

    for column_name in HIDDEN_TRANSACTION_COLUMNS:
        column_letter = get_column_letter(OUTPUT_COLUMNS.index(column_name) + 1)
        transactions_ws.column_dimensions[column_letter].hidden = True
    categories_ws.sheet_state = "hidden"
    accounts_ws.sheet_state = "hidden"
    visible_sheets = [
        overview_ws,
        needs_review_ws,
        categories_budget_ws,
        monthly_summary_ws,
        income_summary_ws,
        cash_flow_summary_ws,
        transactions_ws,
        venmo_activity_ws,
        reconciliation_ws,
    ]
    if income_routing_ws is not None:
        visible_sheets.append(income_routing_ws)
    workbook._sheets = visible_sheets + [
        categories_ws,
        accounts_ws,
    ]
    workbook.active = 0
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
