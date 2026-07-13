from __future__ import annotations

from datetime import datetime

from accounting_pipeline.config import get_owner_buckets
from openpyxl.utils import get_column_letter


SPENDING_EXCLUDED_CATEGORIES = {"Financial – Credit Card Payment"}
SPENDING_EXCLUDED_MAIN_CATEGORIES = {"Transfers", "Income"}
INCOME_MAIN_CATEGORY = "Income"
INCOME_PAYCHECK_PREFIX = "Income – Paycheck:"
INCOME_OTHER_SOURCES_LABEL = "Income – Other Sources"
BUDGET_SHEET_NAME = "'Categories & Budget'"
BUDGET_TARGET_COLUMN = "G"
BUDGET_TARGET_TYPE_COLUMN = "H"


def get_budget_target_formula(label_cell: str) -> str:
    """Return a formula that looks up an editable monthly target by summary label."""
    return (
        f'=IFERROR(SUMIFS({BUDGET_SHEET_NAME}!${BUDGET_TARGET_COLUMN}:${BUDGET_TARGET_COLUMN},'
        f'{BUDGET_SHEET_NAME}!$A:$A,{label_cell}),"")'
    )


def get_budget_target_type_formula(label_cell: str) -> str:
    """Return a formula that looks up the target behavior for a summary label."""
    return (
        f'=IFERROR(INDEX({BUDGET_SHEET_NAME}!${BUDGET_TARGET_TYPE_COLUMN}:${BUDGET_TARGET_TYPE_COLUMN},'
        f'MATCH({label_cell},{BUDGET_SHEET_NAME}!$A:$A,0)),"")'
    )


def get_budget_variance_formula(avg_cell: str, target_cell: str, target_type_cell: str) -> str:
    """Return a formula where positive variance means meeting or beating target."""
    return (
        f'=IF(OR({target_cell}="",{target_type_cell}=""),"",'
        f'IF({target_type_cell}="min",{avg_cell}-{target_cell},'
        f'IF({target_type_cell}="max",{target_cell}-{avg_cell},'
        f'IF({target_type_cell}="exact",-ABS({avg_cell}-{target_cell}),""))))'
    )


def get_ytd_columns(reporting_months: list[tuple[datetime, datetime, str]]) -> list[int]:
    """Return worksheet columns in the latest loaded calendar year."""
    if not reporting_months:
        return []
    latest_year = reporting_months[-1][0].year
    return [
        column_number
        for column_number, (month_start, _month_end, _sheet_name) in enumerate(reporting_months, start=2)
        if month_start.year == latest_year
    ]


def populate_monthly_ytd_budget_cells(
    worksheet,
    row_number: int,
    total_column: int,
    target_column: int,
    target_type_column: int,
    monthly_variance_column: int,
    ytd_actual_column: int,
    ytd_target_column: int,
    ytd_variance_column: int,
    average_column: int,
    average_variance_column: int,
    reporting_month_count: int,
    ytd_columns: list[int],
) -> None:
    """Add latest-month, year-to-date, and average budget comparison formulas."""
    latest_month_letter = get_column_letter(total_column - 1)
    target_letter = get_column_letter(target_column)
    target_type_letter = get_column_letter(target_type_column)
    ytd_actual_letter = get_column_letter(ytd_actual_column)
    ytd_target_letter = get_column_letter(ytd_target_column)
    total_letter = get_column_letter(total_column)
    average_letter = get_column_letter(average_column)
    ytd_cells = ",".join(f"{get_column_letter(column_number)}{row_number}" for column_number in ytd_columns)
    ytd_month_count = max(len(ytd_columns), 1)

    worksheet.cell(row=row_number, column=target_column).value = get_budget_target_formula(f"$A{row_number}")
    worksheet.cell(row=row_number, column=target_type_column).value = get_budget_target_type_formula(f"$A{row_number}")
    worksheet.cell(row=row_number, column=monthly_variance_column).value = get_budget_variance_formula(
        f"{latest_month_letter}{row_number}",
        f"{target_letter}{row_number}",
        f"{target_type_letter}{row_number}",
    )
    worksheet.cell(row=row_number, column=ytd_actual_column).value = (
        f"=SUM({ytd_cells})" if ytd_cells else ""
    )
    worksheet.cell(row=row_number, column=ytd_target_column).value = (
        f'=IF({target_letter}{row_number}="","",{target_letter}{row_number}*{ytd_month_count})'
    )
    worksheet.cell(row=row_number, column=ytd_variance_column).value = get_budget_variance_formula(
        f"{ytd_actual_letter}{row_number}",
        f"{ytd_target_letter}{row_number}",
        f"{target_type_letter}{row_number}",
    )
    worksheet.cell(row=row_number, column=average_column).value = (
        f'=IF({total_letter}{row_number}="","",{total_letter}{row_number}/{max(reporting_month_count, 1)})'
    )
    worksheet.cell(row=row_number, column=average_variance_column).value = get_budget_variance_formula(
        f"{average_letter}{row_number}",
        f"{target_letter}{row_number}",
        f"{target_type_letter}{row_number}",
    )


def get_reporting_months(post_dates: list[datetime]) -> list[tuple[datetime, datetime, str]]:
    """Return sorted calendar-month periods present in the transaction data."""
    months: dict[tuple[int, int], tuple[datetime, datetime, str]] = {}

    for post_date in post_dates:
        key = (post_date.year, post_date.month)
        if key in months:
            continue

        month_start = datetime(post_date.year, post_date.month, 1)
        if post_date.month == 12:
            next_month = datetime(post_date.year + 1, 1, 1)
        else:
            next_month = datetime(post_date.year, post_date.month + 1, 1)
        month_end = next_month.replace(day=1)  # temporary for timedelta-free rollover
        month_end = datetime.fromordinal(month_end.toordinal() - 1)
        months[key] = (month_start, month_end, f"{post_date:%Y-%m} Monthly")

    return [months[key] for key in sorted(months)]


def get_summary_categories(category_rows: list[dict[str, str]]) -> list[str]:
    """Return the leaf categories included in spend/income summaries."""
    return [
        category_row["combined"]
        for category_row in category_rows
        if category_row["main_category"] not in SPENDING_EXCLUDED_MAIN_CATEGORIES
        and category_row["combined"] not in SPENDING_EXCLUDED_CATEGORIES
    ]


def get_summary_category_groups(category_rows: list[dict[str, str]]) -> list[tuple[str, list[str]]]:
    """Return summary categories grouped by main category, preserving template order."""
    grouped_categories: list[tuple[str, list[str]]] = []
    categories_by_main: dict[str, list[str]] = {}

    for category_row in category_rows:
        main_category = category_row["main_category"]
        combined = category_row["combined"]
        if main_category in SPENDING_EXCLUDED_MAIN_CATEGORIES or combined in SPENDING_EXCLUDED_CATEGORIES:
            continue
        if main_category not in categories_by_main:
            categories_by_main[main_category] = []
            grouped_categories.append((main_category, categories_by_main[main_category]))
        categories_by_main[main_category].append(combined)

    return grouped_categories


def get_income_category_groups(category_rows: list[dict[str, str]]) -> list[tuple[str, list[str]]]:
    """Return income categories grouped by main category, preserving template order."""
    grouped_categories: list[tuple[str, list[str]]] = []
    categories_by_summary_label: dict[str, list[str]] = {}

    for category_row in category_rows:
        if category_row["main_category"] != INCOME_MAIN_CATEGORY:
            continue

        combined = category_row["combined"]
        if combined.startswith(INCOME_PAYCHECK_PREFIX):
            summary_label = "Income – Paycheck"
        else:
            summary_label = INCOME_OTHER_SOURCES_LABEL

        if summary_label not in categories_by_summary_label:
            categories_by_summary_label[summary_label] = []
            grouped_categories.append((summary_label, categories_by_summary_label[summary_label]))
        categories_by_summary_label[summary_label].append(combined)

    return grouped_categories


def populate_monthly_summary(
    worksheet,
    category_groups: list[tuple[str, list[str]]],
    reporting_months: list[tuple[datetime, datetime, str]],
    amount_range: str,
    category_range: str,
    bucket_range: str,
    post_date_range: str,
    owner_buckets: list[str] | None = None,
    credit_bucket: str = "Credit",
    start_row: int = 1,
) -> None:
    """Write one review-friendly monthly summary sheet."""
    month_start_row = start_row + 1
    month_end_row = start_row + 2
    ytd_columns = get_ytd_columns(reporting_months)
    worksheet.cell(start_row, 1).value = "Category"
    worksheet.cell(month_start_row, 1).value = "month_start"
    worksheet.cell(month_end_row, 1).value = "month_end"
    for column_number, (month_start, month_end, _sheet_name) in enumerate(reporting_months, start=2):
        worksheet.cell(row=start_row, column=column_number).value = f"{month_start:%b %Y} spend"
        worksheet.cell(row=month_start_row, column=column_number).value = month_start
        worksheet.cell(row=month_end_row, column=column_number).value = month_end
    total_column = len(reporting_months) + 2
    worksheet.cell(row=start_row, column=total_column).value = "Total"
    target_column = total_column + 1
    target_type_column = total_column + 2
    monthly_variance_column = total_column + 3
    ytd_actual_column = total_column + 4
    ytd_target_column = total_column + 5
    ytd_variance_column = total_column + 6
    average_column = total_column + 7
    average_variance_column = total_column + 8
    worksheet.cell(row=start_row, column=target_column).value = "Monthly target"
    worksheet.cell(row=start_row, column=target_type_column).value = "Target type"
    worksheet.cell(row=start_row, column=monthly_variance_column).value = "Monthly variance"
    worksheet.cell(row=start_row, column=ytd_actual_column).value = "YTD actual"
    worksheet.cell(row=start_row, column=ytd_target_column).value = "YTD target"
    worksheet.cell(row=start_row, column=ytd_variance_column).value = "YTD variance"
    worksheet.cell(row=start_row, column=average_column).value = "Monthly average"
    worksheet.cell(row=start_row, column=average_variance_column).value = "Average variance"

    def populate_spending_budget_cells(row_number: int) -> None:
        populate_monthly_ytd_budget_cells(
            worksheet,
            row_number,
            total_column,
            target_column,
            target_type_column,
            monthly_variance_column,
            ytd_actual_column,
            ytd_target_column,
            ytd_variance_column,
            average_column,
            average_variance_column,
            max(len(reporting_months), 1),
            ytd_columns,
        )

    row_number = start_row + 3
    for main_category, category_names in category_groups:
        total_row_number = row_number
        worksheet[f"A{total_row_number}"] = main_category
        row_number += 1

        child_rows = []
        for category_name in category_names:
            worksheet[f"A{row_number}"] = category_name
            child_rows.append(row_number)
            month_cells = []
            for column_number in range(2, total_column):
                column_letter = get_column_letter(column_number)
                month_cells.append(f"{column_letter}{row_number}")
                worksheet.cell(row=row_number, column=column_number).value = (
                    f'=IF($A{row_number}="","",-SUMIFS({amount_range},{category_range},$A{row_number},'
                    f'{post_date_range},">="&{column_letter}${month_start_row},'
                    f'{post_date_range},"<="&{column_letter}${month_end_row}))'
            )
            worksheet.cell(row=row_number, column=total_column).value = f"=SUM({','.join(month_cells)})"
            populate_spending_budget_cells(row_number)
            row_number += 1

        for column_number in range(2, total_column + 1):
            column_letter = get_column_letter(column_number)
            child_cells = ",".join(f"{column_letter}{child_row}" for child_row in child_rows)
            worksheet.cell(row=total_row_number, column=column_number).value = f"=SUM({child_cells})"
        populate_spending_budget_cells(total_row_number)

    latest_month_column = total_column - 1
    compact_ytd_column = target_column
    compact_average_column = monthly_variance_column
    bucket_start_row = row_number + 1
    worksheet[f"A{bucket_start_row}"] = "Owner bucket summary"
    worksheet.cell(row=bucket_start_row, column=latest_month_column).value = "Latest month"
    worksheet.cell(row=bucket_start_row, column=compact_ytd_column).value = "YTD actual"
    worksheet.cell(row=bucket_start_row, column=compact_average_column).value = "Monthly average"

    for row_number, owner_bucket in enumerate(
        get_spending_owner_buckets(owner_buckets, credit_bucket),
        start=bucket_start_row + 1,
    ):
        worksheet[f"A{row_number}"] = owner_bucket
        month_cells = []
        for column_number in range(2, total_column):
            column_letter = get_column_letter(column_number)
            month_cells.append(f"{column_letter}{row_number}")
            worksheet.cell(row=row_number, column=column_number).value = (
                f'=IF($A{row_number}="","",-SUMIFS({amount_range},{bucket_range},$A{row_number},'
                f'{post_date_range},">="&{column_letter}${month_start_row},'
                f'{post_date_range},"<="&{column_letter}${month_end_row},'
                f'{category_range},"<>Transfers*",{category_range},"<>Income*",'
                f'{category_range},"<>Financial – Credit Card Payment"))'
            )
        worksheet.cell(row=row_number, column=compact_ytd_column).value = f"=SUM({','.join(month_cells)})"
        worksheet.cell(row=row_number, column=compact_average_column).value = (
            f'=IF({get_column_letter(compact_ytd_column)}{row_number}="","",'
            f'{get_column_letter(compact_ytd_column)}{row_number}/{max(len(reporting_months), 1)})'
        )


def populate_income_summary(
    worksheet,
    category_groups: list[tuple[str, list[str]]],
    reporting_months: list[tuple[datetime, datetime, str]],
    amount_range: str,
    category_range: str,
    account_type_range: str,
    post_date_range: str,
    start_row: int = 1,
) -> None:
    """Write a monthly income summary sheet."""
    month_start_row = start_row + 1
    month_end_row = start_row + 2
    ytd_columns = get_ytd_columns(reporting_months)
    worksheet.cell(start_row, 1).value = "Income source"
    worksheet.cell(month_start_row, 1).value = "month_start"
    worksheet.cell(month_end_row, 1).value = "month_end"
    for column_number, (month_start, month_end, _sheet_name) in enumerate(reporting_months, start=2):
        worksheet.cell(row=start_row, column=column_number).value = f"{month_start:%b %Y} income"
        worksheet.cell(row=month_start_row, column=column_number).value = month_start
        worksheet.cell(row=month_end_row, column=column_number).value = month_end
    total_column = len(reporting_months) + 2
    worksheet.cell(row=start_row, column=total_column).value = "Total"
    target_column = total_column + 1
    target_type_column = total_column + 2
    monthly_variance_column = total_column + 3
    ytd_actual_column = total_column + 4
    ytd_target_column = total_column + 5
    ytd_variance_column = total_column + 6
    average_column = total_column + 7
    average_variance_column = total_column + 8
    worksheet.cell(row=start_row, column=target_column).value = "Monthly target"
    worksheet.cell(row=start_row, column=target_type_column).value = "Target type"
    worksheet.cell(row=start_row, column=monthly_variance_column).value = "Monthly variance"
    worksheet.cell(row=start_row, column=ytd_actual_column).value = "YTD actual"
    worksheet.cell(row=start_row, column=ytd_target_column).value = "YTD target"
    worksheet.cell(row=start_row, column=ytd_variance_column).value = "YTD variance"
    worksheet.cell(row=start_row, column=average_column).value = "Monthly average"
    worksheet.cell(row=start_row, column=average_variance_column).value = "Average variance"

    def populate_income_budget_cells(row_number: int) -> None:
        populate_monthly_ytd_budget_cells(
            worksheet,
            row_number,
            total_column,
            target_column,
            target_type_column,
            monthly_variance_column,
            ytd_actual_column,
            ytd_target_column,
            ytd_variance_column,
            average_column,
            average_variance_column,
            max(len(reporting_months), 1),
            ytd_columns,
        )

    row_number = start_row + 3
    for summary_label, category_names in category_groups:
        total_row_number = row_number
        worksheet[f"A{total_row_number}"] = summary_label
        row_number += 1

        child_rows = []
        for category_name in category_names:
            worksheet[f"A{row_number}"] = category_name
            child_rows.append(row_number)
            month_cells = []
            for column_number in range(2, total_column):
                column_letter = get_column_letter(column_number)
                month_cells.append(f"{column_letter}{row_number}")
                worksheet.cell(row=row_number, column=column_number).value = (
                    f'=IF($A{row_number}="","",SUMIFS({amount_range},{category_range},$A{row_number},'
                    f'{post_date_range},">="&{column_letter}${month_start_row},'
                    f'{post_date_range},"<="&{column_letter}${month_end_row},'
                    f'{account_type_range},"<>credit_card",{amount_range},">0"))'
            )
            worksheet.cell(row=row_number, column=total_column).value = f"=SUM({','.join(month_cells)})"
            populate_income_budget_cells(row_number)
            row_number += 1

        for column_number in range(2, total_column + 1):
            column_letter = get_column_letter(column_number)
            child_cells = ",".join(f"{column_letter}{child_row}" for child_row in child_rows)
            worksheet.cell(row=total_row_number, column=column_number).value = f"=SUM({child_cells})"
        populate_income_budget_cells(total_row_number)
        target_letter = get_column_letter(target_column)
        target_type_letter = get_column_letter(target_type_column)
        child_target_cells = ",".join(f"{target_letter}{child_row}" for child_row in child_rows)
        worksheet.cell(row=total_row_number, column=target_column).value = (
            f'=IF(COUNT({child_target_cells})=0,"",SUM({child_target_cells}))'
        )
        worksheet.cell(row=total_row_number, column=target_type_column).value = (
            f'=IF({target_letter}{total_row_number}="","","min")'
        )


def get_spending_owner_buckets(
    owner_buckets: list[str] | None = None,
    credit_bucket: str = "Credit",
) -> list[str]:
    """Return owner buckets that should be summarized as personal/family spending."""
    buckets = owner_buckets if owner_buckets is not None else get_owner_buckets()
    return [bucket for bucket in buckets if bucket != credit_bucket]


def populate_cash_flow_summary(
    worksheet,
    reporting_months: list[tuple[datetime, datetime, str]],
    amount_range: str,
    category_range: str,
    account_type_range: str,
    post_date_range: str,
    start_row: int = 1,
) -> None:
    """Write a monthly cash-flow summary that separates spending from transfers."""
    month_start_row = start_row + 1
    month_end_row = start_row + 2
    worksheet.cell(start_row, 1).value = "Cash flow"
    worksheet.cell(month_start_row, 1).value = "month_start"
    worksheet.cell(month_end_row, 1).value = "month_end"
    for column_number, (month_start, month_end, _sheet_name) in enumerate(reporting_months, start=2):
        worksheet.cell(row=start_row, column=column_number).value = f"{month_start:%b %Y}"
        worksheet.cell(row=month_start_row, column=column_number).value = month_start
        worksheet.cell(row=month_end_row, column=column_number).value = month_end

    total_column = len(reporting_months) + 2
    worksheet.cell(row=start_row, column=total_column).value = "Total"
    ytd_columns = get_ytd_columns(reporting_months)
    ytd_column = total_column + 1
    worksheet.cell(row=start_row, column=ytd_column).value = "YTD"
    latest_year = reporting_months[-1][0].year if reporting_months else None
    year_columns: dict[int, list[int]] = {}
    for column_number, (month_start, _month_end, _sheet_name) in enumerate(reporting_months, start=2):
        year_columns.setdefault(month_start.year, []).append(column_number)
    first_year_summary_column = ytd_column + 1
    for offset, year in enumerate(sorted(year_columns)):
        header = f"{year} YTD" if year == latest_year else f"{year} total"
        worksheet.cell(row=start_row, column=first_year_summary_column + offset).value = header

    row_labels = [
        "Cash in",
        "Cash out - spending",
        "Net external cash flow",
        "",
        "Excluded cash out from spending",
        "Internal transfers out",
        "Credit card payments out",
        "Other transfers out",
        "",
        "Needs review",
        "Uncategorized - Needs Review cash out",
    ]

    first_data_row = start_row + 3
    for row_number, label in enumerate(row_labels, start=first_data_row):
        worksheet[f"A{row_number}"] = label

    cash_in_row = first_data_row
    cash_out_row = first_data_row + 1
    net_cash_flow_row = first_data_row + 2
    internal_transfer_row = first_data_row + 5
    card_payment_row = first_data_row + 6
    other_transfer_row = first_data_row + 7
    uncategorized_row = first_data_row + 10

    for column_number in range(2, total_column):
        column_letter = get_column_letter(column_number)
        month_filter = (
            f'{post_date_range},">="&{column_letter}${month_start_row},'
            f'{post_date_range},"<="&{column_letter}${month_end_row},'
            f'{account_type_range},"<>credit_card"'
        )
        worksheet.cell(row=cash_in_row, column=column_number).value = (
            f'=SUMIFS({amount_range},{month_filter},{amount_range},">0",'
            f'{category_range},"<>Transfers*",{category_range},"<>Financial – Credit Card Payment")'
        )
        worksheet.cell(row=cash_out_row, column=column_number).value = (
            f'=-SUMIFS({amount_range},{month_filter},{amount_range},"<0",'
            f'{category_range},"<>Transfers*",{category_range},"<>Financial – Credit Card Payment")'
        )
        worksheet.cell(row=net_cash_flow_row, column=column_number).value = (
            f"={column_letter}{cash_in_row}-{column_letter}{cash_out_row}"
        )
        worksheet.cell(row=internal_transfer_row, column=column_number).value = (
            f'=-SUMIFS({amount_range},{month_filter},{amount_range},"<0",'
            f'{category_range},"Transfers – Internal Transfer")'
        )
        worksheet.cell(row=card_payment_row, column=column_number).value = (
            f'=-SUMIFS({amount_range},{month_filter},{amount_range},"<0",'
            f'{category_range},"Financial – Credit Card Payment")'
        )
        worksheet.cell(row=other_transfer_row, column=column_number).value = (
            f'=-SUMIFS({amount_range},{month_filter},{amount_range},"<0",'
            f'{category_range},"Transfers*",{category_range},"<>Transfers – Internal Transfer")'
        )
        worksheet.cell(row=uncategorized_row, column=column_number).value = (
            f'=-SUMIFS({amount_range},{month_filter},{amount_range},"<0",'
            f'{category_range},"Uncategorized – Needs Review")'
        )

    for column_number in range(2, total_column + 1):
        if column_number == total_column:
            last_month_column = get_column_letter(total_column - 1)
            for row_number in range(first_data_row, first_data_row + len(row_labels)):
                worksheet.cell(row=row_number, column=column_number).value = (
                    ""
                    if worksheet[f"A{row_number}"].value == ""
                    else f"=SUM(B{row_number}:{last_month_column}{row_number})"
                )
    ytd_cells_by_row = {
        row_number: ",".join(f"{get_column_letter(column_number)}{row_number}" for column_number in ytd_columns)
        for row_number in range(first_data_row, first_data_row + len(row_labels))
    }
    for row_number, ytd_cells in ytd_cells_by_row.items():
        worksheet.cell(row=row_number, column=ytd_column).value = (
            ""
            if worksheet[f"A{row_number}"].value == ""
            else f"=SUM({ytd_cells})" if ytd_cells else ""
        )
    for offset, year in enumerate(sorted(year_columns)):
        year_summary_column = first_year_summary_column + offset
        for row_number in range(first_data_row, first_data_row + len(row_labels)):
            year_cells = ",".join(
                f"{get_column_letter(column_number)}{row_number}"
                for column_number in year_columns[year]
            )
            worksheet.cell(row=row_number, column=year_summary_column).value = (
                ""
                if worksheet[f"A{row_number}"].value == ""
                else f"=SUM({year_cells})"
            )
