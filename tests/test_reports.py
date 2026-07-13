from __future__ import annotations

import sys
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from accounting_pipeline.reports.monthly_summary import (
    get_income_category_groups,
    get_spending_owner_buckets,
    get_reporting_months,
    get_summary_category_groups,
    get_summary_categories,
    populate_cash_flow_summary,
    populate_income_summary,
    populate_monthly_summary,
)
from accounting_pipeline.reports.reconciliation import build_reconciliation_rows
from accounting_pipeline.models import Account, StatementMetadata


TEST_ACCOUNTS = [
    Account("1001", "Checking 1001", "checking", "Personal", "bank", "1001"),
    Account("2002", "Savings 2002", "savings", "Family", "bank", "2002"),
    Account("3003", "Credit Card 3003", "credit_card", "Credit", "card", "3003"),
    Account("4004", "Checking 4004", "checking", "Personal", "bank", "4004"),
    Account("5005", "Checking 5005", "checking", "Family", "bank", "5005"),
]


class ReconciliationReportTests(unittest.TestCase):
    def test_build_reconciliation_rows_uses_statement_metadata_and_formulas(self) -> None:
        statement_metadata = {
            "3003": [
                StatementMetadata(
                    start_date=datetime(2026, 2, 26),
                    end_date=datetime(2026, 3, 25),
                    opening_balance=135.70,
                    closing_balance=302.30,
                )
            ]
        }

        rows = build_reconciliation_rows(
            statement_metadata,
            amount_range="transactions!$I$2:$I$171",
            account_range="transactions!$A$2:$A$171",
            post_date_range="transactions!$G$2:$G$171",
            accounts=TEST_ACCOUNTS,
        )

        self.assertEqual(len(rows), 5)
        credit_row = next(row for row in rows if row[0] == "3003")

        self.assertEqual(credit_row[3], datetime(2026, 2, 26))
        self.assertEqual(credit_row[4], datetime(2026, 3, 25))
        self.assertEqual(credit_row[5], 135.70)
        self.assertEqual(credit_row[6], 302.30)
        self.assertIn('SUMIFS(transactions!$I$2:$I$171,transactions!$A$2:$A$171,A6', credit_row[7])
        self.assertEqual(
            credit_row[8],
            '=IF(OR(F6="",H6=""),"",IF(C6="credit_card",F6-H6,F6+H6))',
        )
        self.assertEqual(
            credit_row[9],
            '=IF(OR(G6="",I6=""),"",G6-I6)',
        )
        self.assertEqual(
            credit_row[10],
            '=IF(J6="","",IF(ABS(J6)<0.01,"OK","Review"))',
        )

    def test_build_reconciliation_rows_leaves_missing_statement_values_blank(self) -> None:
        rows = build_reconciliation_rows(
            {},
            amount_range="transactions!$I$2:$I$171",
            account_range="transactions!$A$2:$A$171",
            post_date_range="transactions!$G$2:$G$171",
            accounts=TEST_ACCOUNTS,
        )

        first_row = rows[0]
        self.assertIsNone(first_row[3])
        self.assertIsNone(first_row[4])
        self.assertIsNone(first_row[5])
        self.assertIsNone(first_row[6])

    def test_build_reconciliation_rows_sorts_by_statement_date(self) -> None:
        statement_metadata = {
            "5005": [
                StatementMetadata(
                    start_date=datetime(2025, 1, 1),
                    end_date=datetime(2025, 1, 31),
                    opening_balance=1.0,
                    closing_balance=2.0,
                )
            ],
            "3003": [
                StatementMetadata(
                    start_date=datetime(2024, 12, 1),
                    end_date=datetime(2024, 12, 31),
                    opening_balance=3.0,
                    closing_balance=4.0,
                )
            ],
        }

        rows = build_reconciliation_rows(
            statement_metadata,
            amount_range="transactions!$I$2:$I$171",
            account_range="transactions!$A$2:$A$171",
            post_date_range="transactions!$G$2:$G$171",
            accounts=TEST_ACCOUNTS,
        )

        first_reported_row = next(row for row in rows if row[3] is not None)
        second_reported_row = next(row for row in rows if row[3] is not None and row is not first_reported_row)

        self.assertEqual(first_reported_row[0], "3003")
        self.assertEqual(first_reported_row[3], datetime(2024, 12, 1))
        self.assertEqual(second_reported_row[0], "5005")
        self.assertEqual(second_reported_row[3], datetime(2025, 1, 1))


class MonthlySummaryReportTests(unittest.TestCase):
    def test_get_reporting_months_returns_unique_sorted_calendar_months(self) -> None:
        months = get_reporting_months(
            [
                datetime(2026, 3, 15),
                datetime(2026, 2, 28),
                datetime(2026, 3, 1),
                datetime(2026, 2, 1),
            ]
        )

        self.assertEqual(
            months,
            [
                (datetime(2026, 2, 1), datetime(2026, 2, 28), "2026-02 Monthly"),
                (datetime(2026, 3, 1), datetime(2026, 3, 31), "2026-03 Monthly"),
            ],
        )

    def test_get_summary_categories_excludes_transfers_income_and_card_payments(self) -> None:
        category_rows = [
            {"main_category": "Food", "combined": "Food – Groceries"},
            {"main_category": "Transfers", "combined": "Transfers – Internal Transfer"},
            {"main_category": "Income", "combined": "Income – Interest"},
            {"main_category": "Financial", "combined": "Financial – Credit Card Payment"},
            {"main_category": "Entertainment", "combined": "Entertainment – Streaming"},
        ]

        categories = get_summary_categories(category_rows)

        self.assertEqual(categories, ["Food – Groceries", "Entertainment – Streaming"])

    def test_get_summary_category_groups_preserves_main_category_totals(self) -> None:
        category_rows = [
            {"main_category": "Food", "combined": "Food – Groceries"},
            {"main_category": "Food", "combined": "Food – Dining Out"},
            {"main_category": "Transfers", "combined": "Transfers – Internal Transfer"},
            {"main_category": "Financial", "combined": "Financial – Credit Card Payment"},
            {"main_category": "Financial", "combined": "Financial – Bank Fees"},
        ]

        groups = get_summary_category_groups(category_rows)

        self.assertEqual(
            groups,
            [
                ("Food", ["Food – Groceries", "Food – Dining Out"]),
                ("Financial", ["Financial – Bank Fees"]),
            ],
        )

    def test_get_income_category_groups_includes_only_income_categories(self) -> None:
        category_rows = [
            {"main_category": "Food", "combined": "Food – Groceries"},
            {"main_category": "Income", "combined": "Income – Paycheck: Riley Payroll"},
            {"main_category": "Income", "combined": "Income – Paycheck: Morgan Payroll"},
            {"main_category": "Income", "combined": "Income – Unemployment"},
            {"main_category": "Income", "combined": "Income – Interest"},
            {"main_category": "Transfers", "combined": "Transfers – Internal Transfer"},
        ]

        groups = get_income_category_groups(category_rows)

        self.assertEqual(
            groups,
            [
                (
                    "Income – Paycheck",
                    ["Income – Paycheck: Riley Payroll", "Income – Paycheck: Morgan Payroll"],
                ),
                ("Income – Other Sources", ["Income – Unemployment", "Income – Interest"]),
            ],
        )

    def test_populate_monthly_summary_writes_category_and_bucket_totals(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        populate_monthly_summary(
            worksheet,
            category_groups=[
                ("Food", ["Food – Groceries", "Food – Dining Out"]),
                ("Entertainment", ["Entertainment – Streaming"]),
            ],
            reporting_months=[
                (datetime(2026, 2, 1), datetime(2026, 2, 28), "2026-02 Monthly"),
                (datetime(2026, 3, 1), datetime(2026, 3, 31), "2026-03 Monthly"),
            ],
            amount_range="transactions!$J$2:$J$171",
            category_range="transactions!$N$2:$N$171",
            bucket_range="transactions!$D$2:$D$171",
            post_date_range="transactions!$G$2:$G$171",
            owner_buckets=["Family", "Personal", "Credit", "Needs Review"],
        )

        self.assertEqual(worksheet["A1"].value, "Category")
        self.assertEqual(worksheet["B1"].value, "Feb 2026 spend")
        self.assertEqual(worksheet["C1"].value, "Mar 2026 spend")
        self.assertEqual(worksheet["D1"].value, "Total")
        self.assertEqual(worksheet["E1"].value, "Monthly target")
        self.assertEqual(worksheet["F1"].value, "Target type")
        self.assertEqual(worksheet["G1"].value, "Monthly variance")
        self.assertEqual(worksheet["H1"].value, "YTD actual")
        self.assertEqual(worksheet["I1"].value, "YTD target")
        self.assertEqual(worksheet["J1"].value, "YTD variance")
        self.assertEqual(worksheet["K1"].value, "Monthly average")
        self.assertEqual(worksheet["L1"].value, "Average variance")
        self.assertEqual(worksheet["A4"].value, "Food")
        self.assertEqual(worksheet["B4"].value, "=SUM(B5,B6)")
        self.assertEqual(
            worksheet["E4"].value,
            '=IF(COUNTIFS(\'Categories & Budget\'!$A:$A,$A4,\'Categories & Budget\'!$G:$G,"<>")=0,"",'
            'IFERROR(SUMIFS(\'Categories & Budget\'!$G:$G,\'Categories & Budget\'!$A:$A,$A4),""))',
        )
        self.assertEqual(
            worksheet["F4"].value,
            '=IFERROR(INDEX(\'Categories & Budget\'!$H:$H,MATCH($A4,\'Categories & Budget\'!$A:$A,0)),"")',
        )
        self.assertEqual(
            worksheet["G4"].value,
            '=IF(OR(E4="",F4=""),"",IF(F4="min",C4-E4,IF(F4="max",E4-C4,IF(F4="exact",-ABS(C4-E4),""))))',
        )
        self.assertEqual(worksheet["H4"].value, "=SUM(B4,C4)")
        self.assertEqual(worksheet["I4"].value, '=IF(E4="","",E4*2)')
        self.assertEqual(
            worksheet["J4"].value,
            '=IF(OR(I4="",F4=""),"",IF(F4="min",H4-I4,IF(F4="max",I4-H4,IF(F4="exact",-ABS(H4-I4),""))))',
        )
        self.assertEqual(worksheet["K4"].value, '=IF(D4="","",D4/2)')
        self.assertEqual(
            worksheet["L4"].value,
            '=IF(OR(E4="",F4=""),"",IF(F4="min",K4-E4,IF(F4="max",E4-K4,IF(F4="exact",-ABS(K4-E4),""))))',
        )
        self.assertEqual(worksheet["A5"].value, "Food – Groceries")
        self.assertEqual(
            worksheet["B5"].value,
            '=IF($A5="","",-SUMIFS(transactions!$J$2:$J$171,transactions!$N$2:$N$171,$A5,'
            'transactions!$G$2:$G$171,">="&B$2,transactions!$G$2:$G$171,"<="&B$3))',
        )
        self.assertEqual(worksheet["D5"].value, "=SUM(B5,C5)")
        self.assertEqual(worksheet["A10"].value, "Owner bucket summary")
        self.assertEqual(worksheet["C10"].value, "Latest month")
        self.assertEqual(worksheet["E10"].value, "YTD actual")
        self.assertEqual(worksheet["G10"].value, "Monthly average")
        self.assertEqual(worksheet["A11"].value, "Family")
        self.assertNotIn("Credit", [worksheet[f"A{row_number}"].value for row_number in range(11, 14)])
        self.assertEqual(
            worksheet["B11"].value,
            '=IF($A11="","",-SUMIFS(transactions!$J$2:$J$171,transactions!$D$2:$D$171,$A11,'
            'transactions!$G$2:$G$171,">="&B$2,transactions!$G$2:$G$171,"<="&B$3,'
            'transactions!$N$2:$N$171,"<>Transfers*",transactions!$N$2:$N$171,"<>Income*",'
            'transactions!$N$2:$N$171,"<>Financial – Credit Card Payment"))',
        )
        self.assertEqual(worksheet["E11"].value, "=SUM(B11,C11)")
        self.assertEqual(worksheet["G11"].value, '=IF(E11="","",E11/2)')
        self.assertIsNone(worksheet["F10"].value)

    def test_get_spending_owner_buckets_excludes_credit_bucket(self) -> None:
        buckets = get_spending_owner_buckets(["Family", "Personal", "Credit", "Needs Review"])

        self.assertEqual(buckets, ["Family", "Personal", "Needs Review"])

    def test_populate_income_summary_writes_income_category_totals(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        populate_income_summary(
            worksheet,
            category_groups=[
                ("Income – Paycheck", ["Income – Paycheck: Riley Payroll", "Income – Paycheck: Morgan Payroll"]),
                ("Income – Other Sources", ["Income – Unemployment", "Income – Interest"]),
            ],
            reporting_months=[
                (datetime(2026, 2, 1), datetime(2026, 2, 28), "2026-02 Monthly"),
                (datetime(2026, 3, 1), datetime(2026, 3, 31), "2026-03 Monthly"),
            ],
            amount_range="transactions!$J$2:$J$171",
            category_range="transactions!$N$2:$N$171",
            account_type_range="transactions!$C$2:$C$171",
            post_date_range="transactions!$G$2:$G$171",
        )

        self.assertEqual(worksheet["A1"].value, "Income source")
        self.assertEqual(worksheet["B1"].value, "Feb 2026 income")
        self.assertEqual(worksheet["C1"].value, "Mar 2026 income")
        self.assertEqual(worksheet["D1"].value, "Total")
        self.assertEqual(worksheet["E1"].value, "Monthly target")
        self.assertEqual(worksheet["F1"].value, "Target type")
        self.assertEqual(worksheet["G1"].value, "Monthly variance")
        self.assertEqual(worksheet["H1"].value, "YTD actual")
        self.assertEqual(worksheet["I1"].value, "YTD target")
        self.assertEqual(worksheet["J1"].value, "YTD variance")
        self.assertEqual(worksheet["K1"].value, "Monthly average")
        self.assertEqual(worksheet["L1"].value, "Average variance")
        self.assertEqual(worksheet["A4"].value, "Income – Paycheck")
        self.assertEqual(worksheet["B4"].value, "=SUM(B5,B6)")
        self.assertEqual(
            worksheet["E4"].value,
            '=IF(COUNTIFS(\'Categories & Budget\'!$A:$A,$A4,\'Categories & Budget\'!$G:$G,"<>")>0,'
            'IFERROR(SUMIFS(\'Categories & Budget\'!$G:$G,\'Categories & Budget\'!$A:$A,$A4),""),'
            'IF(COUNT(E5,E6)=0,"",SUM(E5,E6)))',
        )
        self.assertEqual(
            worksheet["F4"].value,
            '=IF(E4="","",IFERROR(INDEX(\'Categories & Budget\'!$H:$H,'
            'MATCH($A4,\'Categories & Budget\'!$A:$A,0)),"min"))',
        )
        self.assertEqual(
            worksheet["G4"].value,
            '=IF(OR(E4="",F4=""),"",IF(F4="min",C4-E4,IF(F4="max",E4-C4,IF(F4="exact",-ABS(C4-E4),""))))',
        )
        self.assertEqual(worksheet["H4"].value, "=SUM(B4,C4)")
        self.assertEqual(worksheet["I4"].value, '=IF(E4="","",E4*2)')
        self.assertEqual(
            worksheet["J4"].value,
            '=IF(OR(I4="",F4=""),"",IF(F4="min",H4-I4,IF(F4="max",I4-H4,IF(F4="exact",-ABS(H4-I4),""))))',
        )
        self.assertEqual(worksheet["K4"].value, '=IF(D4="","",D4/2)')
        self.assertEqual(worksheet["A5"].value, "Income – Paycheck: Riley Payroll")
        self.assertEqual(
            worksheet["B5"].value,
            '=IF($A5="","",SUMIFS(transactions!$J$2:$J$171,transactions!$N$2:$N$171,$A5,'
            'transactions!$G$2:$G$171,">="&B$2,transactions!$G$2:$G$171,"<="&B$3,'
            'transactions!$C$2:$C$171,"<>credit_card",transactions!$J$2:$J$171,">0"))',
        )
        self.assertEqual(worksheet["D5"].value, "=SUM(B5,C5)")

    def test_populate_cash_flow_summary_writes_cash_and_excluded_sections(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        populate_cash_flow_summary(
            worksheet,
            reporting_months=[(datetime(2026, 2, 1), datetime(2026, 2, 28), "2026-02 Monthly")],
            amount_range="transactions!$J$2:$J$171",
            category_range="transactions!$N$2:$N$171",
            account_type_range="transactions!$C$2:$C$171",
            post_date_range="transactions!$G$2:$G$171",
        )

        self.assertEqual(worksheet["A1"].value, "Cash flow")
        self.assertEqual(worksheet["B1"].value, "Feb 2026")
        self.assertEqual(worksheet["C1"].value, "Total")
        self.assertEqual(worksheet["D1"].value, "YTD")
        self.assertEqual(worksheet["E1"].value, "2026 YTD")
        self.assertEqual(worksheet["A4"].value, "Cash in")
        self.assertEqual(
            worksheet["B4"].value,
            '=SUMIFS(transactions!$J$2:$J$171,transactions!$G$2:$G$171,">="&B$2,'
            'transactions!$G$2:$G$171,"<="&B$3,transactions!$C$2:$C$171,"<>credit_card",'
            'transactions!$J$2:$J$171,">0",transactions!$N$2:$N$171,"<>Transfers*",'
            'transactions!$N$2:$N$171,"<>Financial – Credit Card Payment")',
        )
        self.assertEqual(worksheet["A8"].value, "Excluded cash out from spending")
        self.assertEqual(worksheet["A9"].value, "Internal transfers out")
        self.assertEqual(
            worksheet["B10"].value,
            '=-SUMIFS(transactions!$J$2:$J$171,transactions!$G$2:$G$171,">="&B$2,'
            'transactions!$G$2:$G$171,"<="&B$3,transactions!$C$2:$C$171,"<>credit_card",'
            'transactions!$J$2:$J$171,"<0",transactions!$N$2:$N$171,"Financial – Credit Card Payment")',
        )
        self.assertEqual(worksheet["A14"].value, "Uncategorized - Needs Review cash out")
        self.assertEqual(worksheet["C14"].value, "=SUM(B14:B14)")
        self.assertEqual(worksheet["D14"].value, "=SUM(B14)")
        self.assertEqual(worksheet["E14"].value, "=SUM(B14)")
        self.assertIsNone(worksheet["F1"].value)


if __name__ == "__main__":
    unittest.main()
