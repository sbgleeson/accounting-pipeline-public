from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from accounting_pipeline.output.workbook_writer import (
    get_group_main_rows,
    populate_categories_budget_reality_formulas,
    populate_categories_budget_sheet,
    populate_overview_sheet,
)


class WorkbookWriterTests(unittest.TestCase):
    def test_group_main_rows_returns_parent_summary_rows(self) -> None:
        groups = [
            ("Income – Paycheck", ["Income – Paycheck: Demo One", "Income – Paycheck: Demo Two"]),
            ("Income – Other Sources", ["Income – Interest"]),
        ]

        self.assertEqual(get_group_main_rows(groups), [10, 13])

    def test_overview_loaded_period_gets_full_width_hero_position(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active

        populate_overview_sheet(
            worksheet,
            metrics=[
                ("Loaded period", "Jan 01, 2026 – Mar 31, 2026", "Date range represented by loaded transactions."),
                ("Accounts", 3, "Configured accounts included in this workbook."),
                ("Transactions", 52, "Normalized rows after duplicate removal."),
                ("Observed income", 1000, "Positive transactions categorized as income."),
                ("Net spending", 500, "Spending after refunds; transfers and card payments excluded."),
                ("Net external cash flow", 250, "Cash-account inflows less outflows, excluding internal transfers."),
                ("Needs review", 2, "Items consolidated on the Needs Review sheet."),
            ],
            income_routing_enabled=False,
        )

        merged_ranges = {str(merged_range) for merged_range in worksheet.merged_cells.ranges}
        self.assertIn("A5:I5", merged_ranges)
        self.assertIn("A6:I6", merged_ranges)
        self.assertEqual(worksheet["A5"].value, "Loaded period")
        self.assertEqual(worksheet["A6"].value, "Jan 01, 2026 – Mar 31, 2026")
        self.assertEqual(worksheet["A9"].value, "Accounts")
        self.assertEqual(worksheet["D9"].value, "Transactions")
        self.assertEqual(worksheet["G9"].value, "Observed income")
        self.assertEqual(worksheet["A13"].value, "Net spending")
        self.assertEqual(worksheet["D13"].value, "Net external cash flow")
        self.assertEqual(worksheet["G13"].value, "Needs review")
        self.assertIn("A9:C9", merged_ranges)
        self.assertIn("D9:F9", merged_ranges)
        self.assertIn("G9:I9", merged_ranges)
        self.assertIn("A13:C13", merged_ranges)
        self.assertIn("D13:F13", merged_ranges)
        self.assertIn("G13:I13", merged_ranges)
        self.assertEqual(worksheet["A19"].value, "Step")
        self.assertEqual(worksheet["B19"].value, "Where to go")
        self.assertEqual(worksheet["E19"].value, "What to check")
        self.assertEqual(worksheet["A20"].value, "1")
        self.assertEqual(worksheet["B20"].value, "Needs Review")
        self.assertIn("B20:D20", merged_ranges)
        self.assertIn("E20:I20", merged_ranges)

    def test_categories_budget_uses_summary_average_as_reality_without_review_status(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        category_rows = [
            {
                "main_category": "Food",
                "sub_category": "Groceries",
                "combined": "Food – Groceries",
            }
        ]
        populate_categories_budget_sheet(
            worksheet,
            category_rows,
            budget_targets=[
                {
                    "budget_label": "Food",
                    "monthly_target": '=ROUND(SUMIF($A:$A,"Income – Paycheck",$G:$G)*0.10,0)',
                    "target_type": "max",
                    "owner_bucket": "Personal",
                    "review_status": "draft",
                    "notes": "Demo food target",
                }
            ],
        )
        populate_categories_budget_reality_formulas(worksheet, "J")

        headers = [worksheet.cell(1, column_number).value for column_number in range(1, worksheet.max_column + 1)]
        self.assertEqual(
            headers,
            [
                "Category / budget label",
                "Budget level",
                "Main category",
                "Subcategory",
                "Combined category",
                "Reality",
                "Monthly target",
                "Target type",
                "Owner bucket",
                "Notes",
            ],
        )
        self.assertNotIn("Review status", headers)
        self.assertEqual(
            worksheet["F2"].value,
            '=IFERROR(INDEX(\'Income Summary\'!$J:$J,MATCH($A2,\'Income Summary\'!$A:$A,0)),'
            'IFERROR(INDEX(\'Spending Summary\'!$J:$J,MATCH($A2,\'Spending Summary\'!$A:$A,0)),""))',
        )
        self.assertEqual(worksheet["J2"].value, "Demo food target")
        self.assertEqual(
            worksheet["G2"].value,
            '=ROUND(SUMIF($A:$A,"Income – Paycheck",$G:$G)*0.10,0)',
        )


if __name__ == "__main__":
    unittest.main()
